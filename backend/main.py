import os
from fastapi import FastAPI, Depends, HTTPException, status, UploadFile, File, Body, Form
from fastapi.responses import Response
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from sqlalchemy import create_engine, text
from sqlalchemy.orm import sessionmaker, Session
from datetime import datetime, timedelta, timezone
from typing import Optional, List
from pydantic import BaseModel
#from passlib.context import CryptContext
from jose import JWTError, jwt
from openpyxl import Workbook, load_workbook
from io import BytesIO
from activity import router as activity_router
from assignments import router as assignments_router
import redis
import re
import bcrypt
from dependencies import (
    get_db, get_current_user, require_admin, require_teacher_or_admin,
    SessionLocal, engine,
    SECRET_KEY, ALGORITHM, ACCESS_TOKEN_EXPIRE_MINUTES,
    oauth2_scheme
)
from activity import router as activity_router
from assignments import router as assignments_router

# Configuration
DATABASE_URL = os.getenv("DATABASE_URL")
if not DATABASE_URL:
    raise RuntimeError("DATABASE_URL environment variable not set")

REDIS_URL = os.getenv("REDIS_URL", "redis://redis:6379")
redis_client = redis.from_url(REDIS_URL, decode_responses=True)
SECRET_KEY = os.getenv("SECRET_KEY")
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 30  # 0.5 hours

# Database setup
engine = create_engine(DATABASE_URL)
SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)

# Redis setup
redis_client = redis.from_url(REDIS_URL, decode_responses=True)

# Password hashing
#pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")

# OAuth2
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="api/token")

# FastAPI app
app = FastAPI(title="Teacher Logbook API", version="1.0.0", root_path="/api")

# 라우터 등록
app.include_router(activity_router, prefix="/api/teacher", tags=["Teacher"])
app.include_router(assignments_router, prefix="/api/assignments", tags=["Assignments"])

# CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # In production, specify exact origins
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# Pydantic Models
class Token(BaseModel):
    access_token: str
    token_type: str
    user_id: str
    role: str
    full_name: Optional[str]


class TokenData(BaseModel):
    user_id: Optional[str] = None


class UserBase(BaseModel):
    user_id: str
    full_name: Optional[str] = None
    role: str


class UserCreate(BaseModel):
    user_id: str
    password: str
    full_name: Optional[str] = None
    role: str
    grade: Optional[int] = None
    class_number: Optional[int] = None
    number_in_class: Optional[int] = None


class UserUpdate(BaseModel):
    full_name: Optional[str] = None
    password: Optional[str] = None


class User(UserBase):
    id: int
    grade: Optional[int] = None
    class_number: Optional[int] = None
    number_in_class: Optional[int] = None
    created_at: datetime

    class Config:
        from_attributes = True


class SubjectCreate(BaseModel):
    subject_name: str
    subject_code: str
    description: Optional[str] = None


class Subject(BaseModel):
    id: int
    subject_name: str
    subject_code: str
    description: Optional[str] = None

    class Config:
        from_attributes = True


class RecordCreate(BaseModel):
    student_user_id: str
    subject_id: int
    content: str


class RecordUpdate(BaseModel):
    content: str


class Record(BaseModel):
    id: int
    student_user_id: str
    subject_id: int
    content: Optional[str] = None
    char_count: int
    byte_count: int
    is_editable_by_student: bool
    created_at: datetime
    updated_at: datetime

    class Config:
        from_attributes = True


class RecordWithDetails(Record):
    student_name: Optional[str] = None
    subject_name: Optional[str] = None
    is_locked: bool = False
    locked_by: Optional[str] = None


class CommentCreate(BaseModel):
    record_id: int
    comment_text: str


class Comment(BaseModel):
    id: int
    record_id: int
    user_id: str
    comment_text: str
    created_at: datetime

    class Config:
        from_attributes = True


class RecordVersion(BaseModel):
    id: int
    record_id: int
    content: Optional[str]
    char_count: int
    byte_count: int
    edited_by: str
    edit_type: str
    created_at: datetime

    class Config:
        from_attributes = True


# Utility functions
def calculate_byte_count(text: str) -> tuple:
    if not text:
        return 0, 0
    char_count = len(text)
    korean_count = len(re.findall(r'[가-힣]', text))
    newline_count = text.count('\n')
    byte_count = korean_count * 3 + (char_count - korean_count - newline_count) + newline_count * 2
    return char_count, byte_count


def verify_password(plain_password: str, hashed_password: str) -> bool:
#    return pwd_context.verify(plain_password, hashed_password)
    try:
        if not plain_password or not hashed_password:
            return False
        # 입력값 검증
        return bcrypt.checkpw(
            plain_password.encode('utf-8'),
            hashed_password.encode('utf-8')
        )
    except Exception as e:
        print(f"verify_password error: {e}")
        return False

def get_password_hash(password: str) -> str:
#    return pwd_context.hash(password)
    try:
        if not password:
            raise ValueError("Password cannot be empty")
        
        password = password.strip()
        
        if len(password.encode('utf-8')) > 72:
            raise ValueError("Password too long (max 72 bytes)")
        
        salt = bcrypt.gensalt(rounds=12)
        hashed = bcrypt.hashpw(password.encode('utf-8'), salt)
        return hashed.decode('utf-8')
    except Exception as e:
        print(f"get_password_hash error: {e}")
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Password hashing failed: {str(e)}"
        )

def create_access_token(data: dict, expires_delta: Optional[timedelta] = None) -> str:
    to_encode = data.copy()
    expire = datetime.now(timezone.utc) + (expires_delta or timedelta(minutes=15))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)


def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


def _build_record_response(r, include_lock=False):
    """Build record response dict"""
    response = {
        "id": r.id,
        "student_user_id": r.student_user_id,
        "subject_id": r.subject_id,
        "content": r.content,
        "char_count": r.char_count,
        "byte_count": r.byte_count,
        "is_editable_by_student": r.is_editable_by_student,
        "created_at": r.created_at,
        "updated_at": r.updated_at,
    }
    if hasattr(r, 'student_name'):
        response.update({"student_name": r.student_name, "subject_name": r.subject_name})
    if include_lock:
        lock_owner = get_lock_owner(r.id)
        response.update({"is_locked": lock_owner is not None, "locked_by": lock_owner})
    return response


def _build_user_response(u):
    """Build user response dict"""
    return {
        "id": u.id,
        "user_id": u.user_id,
        "full_name": u.full_name,
        "role": u.role,
        "grade": u.grade,
        "class_number": u.class_number,
        "number_in_class": u.number_in_class,
        "created_at": u.created_at
    }


# Lock management functions
def acquire_lock(record_id: int, user_id: str, duration_minutes: int = 30) -> bool:
    """Acquire edit lock for a record"""
    lock_key = f"record_lock:{record_id}"
    
    # Check if already locked
    existing_lock = redis_client.get(lock_key)
    if existing_lock and existing_lock != user_id:
        return False
    
    # Set lock with expiration
    redis_client.setex(lock_key, duration_minutes * 60, user_id)
    return True


def release_lock(record_id: int, user_id: str) -> bool:
    """Release edit lock for a record"""
    lock_key = f"record_lock:{record_id}"
    
    # Only release if current user owns the lock
    current_lock = redis_client.get(lock_key)
    if current_lock == user_id:
        redis_client.delete(lock_key)
        return True
    return False


def get_lock_owner(record_id: int) -> Optional[str]:
    """Get the current lock owner for a record"""
    lock_key = f"record_lock:{record_id}"
    return redis_client.get(lock_key)


def extend_lock(record_id: int, user_id: str, duration_minutes: int = 30) -> bool:
    """Extend lock duration"""
    lock_key = f"record_lock:{record_id}"
    current_lock = redis_client.get(lock_key)
    
    if current_lock == user_id:
        redis_client.expire(lock_key, duration_minutes * 60)
        return True
    return False


# API Routes

@app.get("/")
async def root():
    return {"message": "Teacher Logbook API", "version": "1.0.0"}


@app.get("/api/health")
async def health():
    return {"status": "healthy"}


@app.post("/api/token", response_model=Token)
async def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    # Query user
    user = db.execute(
        text("SELECT * FROM users WHERE user_id = :user_id"),
        {"user_id": form_data.username}
    ).fetchone()
    
    if not user or not verify_password(form_data.password, user.password_hash):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Incorrect username or password",
            headers={"WWW-Authenticate": "Bearer"},
        )
    
    access_token_expires = timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    access_token = create_access_token(
        data={"sub": user.user_id}, expires_delta=access_token_expires
    )
    
    return {
        "access_token": access_token,
        "token_type": "bearer",
        "user_id": user.user_id,
        "role": user.role,
        "full_name": user.full_name
    }


@app.get("/api/users/me", response_model=User)
async def read_users_me(current_user = Depends(get_current_user)):
    return {
        "id": current_user.id,
        "user_id": current_user.user_id,
        "full_name": current_user.full_name,
        "role": current_user.role,
        "grade": current_user.grade,
        "class_number": current_user.class_number,
        "number_in_class": current_user.number_in_class,
        "created_at": current_user.created_at
    }


@app.put("/api/users/me", response_model=User)
async def update_user_me(
    user_update: UserUpdate,
    current_user = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    update_data = {}
    if user_update.full_name:
        update_data["full_name"] = user_update.full_name
    if user_update.password:
        update_data["password_hash"] = get_password_hash(user_update.password)
    
    if update_data:
        update_data["updated_at"] = datetime.now(timezone.utc)
        set_clause = ", ".join([f"{k} = :{k}" for k in update_data.keys()])
        query = f"UPDATE users SET {set_clause} WHERE user_id = :user_id RETURNING *"
        update_data["user_id"] = current_user.user_id
        
        result = db.execute(text(query), update_data)
        db.commit()
        updated_user = result.fetchone()
        
        return _build_user_response(updated_user)
    
    return _build_user_response(current_user)


@app.post("/api/token/refresh", response_model=Token)
async def refresh_token(current_user = Depends(get_current_user)):
    access_token_expires = timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    access_token = create_access_token(
        data={"sub": current_user.user_id}, expires_delta=access_token_expires
    )
    
    return {
        "access_token": access_token,
        "token_type": "bearer",
        "user_id": current_user.user_id,
        "role": current_user.role,
        "full_name": current_user.full_name
    }


@app.get("/api/subjects", response_model=List[Subject])
async def get_subjects(
    current_user = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    result = db.execute(text("SELECT * FROM subjects ORDER BY subject_name"))
    subjects = result.fetchall()
    return [
        {
            "id": s.id,
            "subject_name": s.subject_name,
            "subject_code": s.subject_code,
            "description": s.description
        }
        for s in subjects
    ]


@app.post("/api/subjects", response_model=Subject)
async def create_subject(
    subject: SubjectCreate,
    current_user = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    # Only admin can create subjects
    if current_user.role != 'admin':
        raise HTTPException(status_code=403, detail="Only admin can create subjects")
    
    try:
        result = db.execute(
            text("""
                INSERT INTO subjects (subject_name, subject_code, description)
                VALUES (:subject_name, :subject_code, :description)
                RETURNING *
                """),
            {
                "subject_name": subject.subject_name,
                "subject_code": subject.subject_code,
                "description": subject.description
            }
        )
        db.commit()
        new_subject = result.fetchone()
        
        return {
            "id": new_subject.id,
            "subject_name": new_subject.subject_name,
            "subject_code": new_subject.subject_code,
            "description": new_subject.description
        }
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=str(e))


@app.get("/api/records", response_model=List[RecordWithDetails])
async def get_records(
    student_user_id: Optional[str] = None,
    subject_id: Optional[int] = None,
    grade: Optional[int] = None,
    class_number: Optional[int] = None,
    current_user = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    if current_user.role == 'student':
        student_user_id = current_user.user_id

    query = "SELECT r.*, u.full_name as student_name, s.subject_name FROM records r LEFT JOIN users u ON r.student_user_id = u.user_id LEFT JOIN subjects s ON r.subject_id = s.id WHERE 1=1"
    params = {}
    
    if student_user_id:
        query += " AND r.student_user_id = :student_user_id"
        params["student_user_id"] = student_user_id
    if subject_id:
        query += " AND r.subject_id = :subject_id"
        params["subject_id"] = subject_id
    if grade:
        query += " AND u.grade = :grade"
        params["grade"] = grade
    if class_number:
        query += " AND u.class_number = :class_number"
        params["class_number"] = class_number

    query += " ORDER BY u.grade, u.class_number, u.number_in_class, s.subject_name"
    
    records = db.execute(text(query), params).fetchall()
    return [_build_record_response(r, include_lock=True) for r in records]


@app.get("/api/records/{record_id}", response_model=RecordWithDetails)
async def get_record(
    record_id: int,
    current_user = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    result = db.execute(
        text("SELECT r.*, u.full_name as student_name, s.subject_name FROM records r LEFT JOIN users u ON r.student_user_id = u.user_id LEFT JOIN subjects s ON r.subject_id = s.id WHERE r.id = :record_id"),
        {"record_id": record_id}
    )
    record = result.fetchone()
    
    if not record:
        raise HTTPException(status_code=404, detail="Record not found")
    if current_user.role == 'student' and record.student_user_id != current_user.user_id:
        raise HTTPException(status_code=403, detail="Access denied")
    
    return _build_record_response(record, include_lock=True)


@app.post("/api/records", response_model=Record)
async def create_record(
    record: RecordCreate,
    current_user = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    if current_user.role == 'student':
        raise HTTPException(status_code=403, detail="Students cannot create records")
    
    char_count, byte_count = calculate_byte_count(record.content)
    
    try:
        result = db.execute(
            text("INSERT INTO records (student_user_id, subject_id, content, char_count, byte_count) VALUES (:student_user_id, :subject_id, :content, :char_count, :byte_count) RETURNING *"),
            {
                "student_user_id": record.student_user_id,
                "subject_id": record.subject_id,
                "content": record.content,
                "char_count": char_count,
                "byte_count": byte_count
            }
        )
        db.commit()
        new_record = result.fetchone()
        
        db.execute(
            text("INSERT INTO record_versions (record_id, content, char_count, byte_count, edited_by, edit_type) VALUES (:record_id, :content, :char_count, :byte_count, :edited_by, :edit_type)"),
            {
                "record_id": new_record.id,
                "content": record.content,
                "char_count": char_count,
                "byte_count": byte_count,
                "edited_by": current_user.user_id,
                "edit_type": "create"
            }
        )
        db.commit()
        
        return _build_record_response(new_record)
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=str(e))


@app.post("/api/records/{record_id}/lock")
async def lock_record(
    record_id: int,
    current_user = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Acquire lock for editing a record"""
    
    # Check if record exists
    result = db.execute(text("SELECT * FROM records WHERE id = :id"), {"id": record_id})
    record = result.fetchone()
    
    if not record:
        raise HTTPException(status_code=404, detail="Record not found")
    
    # Check permissions
    if current_user.role == 'student':
        if record.student_user_id != current_user.user_id:
            raise HTTPException(status_code=403, detail="Access denied")
        if not record.is_editable_by_student:
            raise HTTPException(status_code=403, detail="Record not editable by student")
    
    # Try to acquire lock
    if acquire_lock(record_id, current_user.user_id):
        return {"message": "Lock acquired", "locked_by": current_user.user_id}
    else:
        lock_owner = get_lock_owner(record_id)
        raise HTTPException(
            status_code=423,
            detail=f"Record is locked by {lock_owner}"
        )


@app.delete("/api/records/{record_id}/lock")
async def unlock_record(record_id: int, current_user = Depends(get_current_user)):
    if release_lock(record_id, current_user.user_id):
        return {"message": "Lock released"}
    raise HTTPException(status_code=400, detail="You don't own this lock")


@app.put("/api/records/{record_id}/lock/extend")
async def extend_record_lock(record_id: int, current_user = Depends(get_current_user)):
    if extend_lock(record_id, current_user.user_id):
        return {"message": "Lock extended"}
    raise HTTPException(status_code=400, detail="You don't own this lock")


@app.put("/api/records/{record_id}", response_model=Record)
async def update_record(
    record_id: int,
    record_update: RecordUpdate,
    current_user = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    result = db.execute(text("SELECT * FROM records WHERE id = :id"), {"id": record_id})
    record = result.fetchone()
    
    if not record:
        raise HTTPException(status_code=404, detail="Record not found")
    
    if current_user.role == 'student':
        if record.student_user_id != current_user.user_id:
            raise HTTPException(status_code=403, detail="Access denied")
        if not record.is_editable_by_student:
            raise HTTPException(status_code=403, detail="Record not editable by student")
    
    lock_owner = get_lock_owner(record_id)
    if lock_owner and lock_owner != current_user.user_id:
        raise HTTPException(status_code=423, detail=f"Record is locked by {lock_owner}")
    
    char_count, byte_count = calculate_byte_count(record_update.content)
    
    try:
        result = db.execute(
            text("UPDATE records SET content = :content, char_count = :char_count, byte_count = :byte_count, updated_at = CURRENT_TIMESTAMP WHERE id = :id RETURNING *"),
            {
                "id": record_id,
                "content": record_update.content,
                "char_count": char_count,
                "byte_count": byte_count
            }
        )
        db.commit()
        updated_record = result.fetchone()
        
        db.execute(
            text("INSERT INTO record_versions (record_id, content, char_count, byte_count, edited_by, edit_type) VALUES (:record_id, :content, :char_count, :byte_count, :edited_by, :edit_type)"),
            {
                "record_id": record_id,
                "content": record_update.content,
                "char_count": char_count,
                "byte_count": byte_count,
                "edited_by": current_user.user_id,
                "edit_type": "update"
            }
        )
        db.commit()
        
        release_lock(record_id, current_user.user_id)
        
        return _build_record_response(updated_record)
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=str(e))


@app.put("/api/records/{record_id}/permissions")
async def update_record_permissions(
    record_id: int,
    is_editable: bool,
    current_user = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    if current_user.role == 'student':
        raise HTTPException(status_code=403, detail="Students cannot change permissions")
    
    try:
        db.execute(text("UPDATE records SET is_editable_by_student = :is_editable, updated_at = CURRENT_TIMESTAMP WHERE id = :id"), {"id": record_id, "is_editable": is_editable})
        db.commit()
        return {"message": "Permissions updated", "is_editable_by_student": is_editable}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=str(e))


@app.get("/api/records/{record_id}/versions", response_model=List[RecordVersion])
async def get_record_versions(
    record_id: int,
    current_user = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    result = db.execute(text("SELECT * FROM records WHERE id = :id"), {"id": record_id})
    record = result.fetchone()
    
    if not record:
        raise HTTPException(status_code=404, detail="Record not found")
    if current_user.role == 'student' and record.student_user_id != current_user.user_id:
        raise HTTPException(status_code=403, detail="Access denied")
    
    result = db.execute(text("SELECT * FROM record_versions WHERE record_id = :record_id ORDER BY created_at DESC"), {"record_id": record_id})
    versions = result.fetchall()
    
    return [
        {
            "id": v.id,
            "record_id": v.record_id,
            "content": v.content,
            "char_count": v.char_count,
            "byte_count": v.byte_count,
            "edited_by": v.edited_by,
            "edit_type": v.edit_type,
            "created_at": v.created_at
        }
        for v in versions
    ]


@app.get("/api/records/{record_id}/comments", response_model=List[Comment])
async def get_record_comments(
    record_id: int,
    current_user = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    result = db.execute(text("SELECT * FROM records WHERE id = :id"), {"id": record_id})
    record = result.fetchone()
    
    if not record:
        raise HTTPException(status_code=404, detail="Record not found")
    if current_user.role == 'student' and record.student_user_id != current_user.user_id:
        raise HTTPException(status_code=403, detail="Access denied")
    
    result = db.execute(text("SELECT * FROM comments WHERE record_id = :record_id ORDER BY created_at DESC"), {"record_id": record_id})
    comments = result.fetchall()
    
    return [{"id": c.id, "record_id": c.record_id, "user_id": c.user_id, "comment_text": c.comment_text, "created_at": c.created_at} for c in comments]


@app.post("/api/records/{record_id}/comments", response_model=Comment)
async def create_comment(
    record_id: int,
    comment: CommentCreate,
    current_user = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    result = db.execute(text("SELECT * FROM records WHERE id = :id"), {"id": record_id})
    record = result.fetchone()
    
    if not record:
        raise HTTPException(status_code=404, detail="Record not found")
    
    if current_user.role == 'student' and record.student_user_id != current_user.user_id:
        raise HTTPException(status_code=403, detail="Access denied")
    
    try:
        result = db.execute(
            text("INSERT INTO comments (record_id, user_id, comment_text) VALUES (:record_id, :user_id, :comment_text) RETURNING *"),
            {"record_id": record_id, "user_id": current_user.user_id, "comment_text": comment.comment_text}
        )
        db.commit()
        new_comment = result.fetchone()
        
        return {"id": new_comment.id, "record_id": new_comment.record_id, "user_id": new_comment.user_id, "comment_text": new_comment.comment_text, "created_at": new_comment.created_at}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=str(e))


# Admin routes
@app.get("/api/admin/users", response_model=List[User])
async def get_all_users(
    current_user = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    if current_user.role != 'admin':
        raise HTTPException(status_code=403, detail="Admin access required")
    
    result = db.execute(text("SELECT * FROM users ORDER BY user_id"))
    users = result.fetchall()
    return [_build_user_response(u) for u in users]


@app.post("/api/admin/users", response_model=User)
async def create_user(
    user: UserCreate,
    current_user = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    if current_user.role != 'admin':
        raise HTTPException(status_code=403, detail="Admin access required")
    
    try:
        password_hash = get_password_hash(user.password)
        
        result = db.execute(
            text("INSERT INTO users (user_id, password_hash, full_name, role, grade, class_number, number_in_class) VALUES (:user_id, :password_hash, :full_name, :role, :grade, :class_number, :number_in_class) RETURNING *"),
            {
                "user_id": user.user_id,
                "password_hash": password_hash,
                "full_name": user.full_name,
                "role": user.role,
                "grade": user.grade,
                "class_number": user.class_number,
                "number_in_class": user.number_in_class
            }
        )
        db.commit()
        new_user = result.fetchone()
        
        return _build_user_response(new_user)
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=str(e))


@app.post("/api/admin/users/bulk-upload")
async def bulk_upload_users(
    users: List[UserCreate],
    current_user = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Bulk upload users (admin only)"""
    
    if current_user.role != 'admin':
        raise HTTPException(status_code=403, detail="Admin access required")
    
    created_count = 0
    errors = []
    
    for user in users:
        try:
            password_hash = get_password_hash(user.password)
            
            db.execute(
                text("""
                    INSERT INTO users (user_id, password_hash, full_name, role, grade, class_number, number_in_class)
                    VALUES (:user_id, :password_hash, :full_name, :role, :grade, :class_number, :number_in_class)
                    """),
                {
                    "user_id": user.user_id,
                    "password_hash": password_hash,
                    "full_name": user.full_name,
                    "role": user.role,
                    "grade": user.grade,
                    "class_number": user.class_number,
                    "number_in_class": user.number_in_class
                }
            )
            db.commit()
            created_count += 1
        except Exception as e:
            db.rollback()
            errors.append({"user_id": user.user_id, "error": str(e)})
    
    return {
        "message": f"Created {created_count} users",
        "created_count": created_count,
        "errors": errors
    }


# ==================== 사용자 비밀번호 초기화 ====================
@app.put("/api/admin/users/{user_id}/reset-password")
def reset_user_password(
    user_id: str,
    request_body: dict = Body(...),
    current_user = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Reset a user's password (admin only)"""
    
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin only")
    
    new_password = request_body.get("new_password")
    if not new_password:
        raise HTTPException(status_code=400, detail="new_password required")
    
    # Raw SQL로 사용자 조회
    result = db.execute(
        text("SELECT * FROM users WHERE user_id = :user_id"),
        {"user_id": user_id}
    )
    user = result.fetchone()
    
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # 비밀번호 해싱
    hashed = bcrypt.hashpw(new_password.encode('utf-8'), bcrypt.gensalt())
    
    # 업데이트
    db.execute(
        text("UPDATE users SET password_hash = :password_hash WHERE user_id = :user_id"),
        {"password_hash": hashed.decode('utf-8'), "user_id": user_id}
    )
    db.commit()
    
    return {"message": f"Password reset for {user_id}"}


# ==================== Excel 템플릿 다운로드 ====================
@app.get("/api/admin/download-template/{template_type}")
def download_excel_template(
    template_type: str,
    current_user = Depends(get_current_user)
):
    """Download Excel template (admin only)"""
    
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin only")
    
    wb = Workbook()
    
    if template_type == "users":
        ws = wb.active
        ws.title = "Users"
        
        # 헤더
        headers = ["user_id", "password", "full_name", "role", "student_number", "grade", "class_number", "number_in_class"]
        ws.append(headers)
        
        # 예시 데이터 (교사)
        ws.append(["T0300", "1234!", "김선생", "teacher", "", "", "", ""])
        ws.append(["T0301", "1234!", "이선생", "teacher", "", "", "", ""])
        
        # 예시 데이터 (학생)
        ws.append(["S20201", "1234!", "홍길동", "student", "20201", "2", "2", "1"])
        ws.append(["S20202", "1234!", "김철수", "student", "20202", "2", "2", "2"])
        
        # 열 너비 조정
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 8
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 15
    
    elif template_type == "subjects":
        ws = wb.active
        ws.title = "Subjects"
        
        # 헤더
        headers = ["subject_code", "subject_name", "description"]
        ws.append(headers)
        
        # 예시 데이터
        ws.append(["KOR", "국어", "국어 과목"])
        ws.append(["ENG", "영어", "영어 과목"])
        ws.append(["MATH", "수학", "수학 과목"])
        
        # 열 너비 조정
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 30
    
    else:
        raise HTTPException(status_code=400, detail="Invalid template type")
    
    # 메모리에 저장
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"{template_type}_template.xlsx"
    
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ==================== Excel 파일 업로드 및 임포트 ====================
@app.post("/api/admin/import-excel/{import_type}")
async def import_excel(
    import_type: str,
    file: UploadFile = File(...),
    current_user = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Import data from Excel file (admin only)"""
    
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin only")
    
    # 파일 읽기
    contents = await file.read()
    wb = load_workbook(BytesIO(contents))
    ws = wb.active
    
    results = {
        "success": 0,
        "failed": 0,
        "errors": []
    }
    
    if import_type == "users":
        # 헤더 건너뛰기
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        
        for idx, row in enumerate(rows, start=2):
            try:
                user_id, password, full_name, role, student_number, grade, class_number, number_in_class = row
                
                # 필수 필드 검증
                if not user_id or not password or not role:
                    results["errors"].append(f"Row {idx}: Missing required fields")
                    results["failed"] += 1
                    continue
                
                # 역할 검증
                if role not in ["admin", "teacher", "student"]:
                    results["errors"].append(f"Row {idx}: Invalid role '{role}'")
                    results["failed"] += 1
                    continue
                
                # 학생인 경우 추가 필드 검증
                if role == "student":
                    if not student_number or not grade or not class_number or not number_in_class:
                        results["errors"].append(f"Row {idx}: Students require student_number, grade, class_number, number_in_class")
                        results["failed"] += 1
                        continue
                
                # 기존 사용자 확인 (Raw SQL)
                result = db.execute(
                    text("SELECT * FROM users WHERE user_id = :user_id"),
                    {"user_id": user_id}
                )
                existing_user = result.fetchone()
                
                if existing_user:
                    # 업데이트
                    hashed = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()) if password else None
                    
                    if hashed:
                        db.execute(
                            text("""UPDATE users 
                                   SET full_name = :full_name, password_hash = :password_hash, role = :role,
                                       student_number = :student_number, grade = :grade, 
                                       class_number = :class_number, number_in_class = :number_in_class
                                   WHERE user_id = :user_id"""),
                            {
                                "user_id": user_id,
                                "full_name": full_name,
                                "password_hash": hashed.decode('utf-8'),
                                "role": role,
                                "student_number": student_number,
                                "grade": grade,
                                "class_number": class_number,
                                "number_in_class": number_in_class
                            }
                        )
                    else:
                        db.execute(
                            text("""UPDATE users 
                                   SET full_name = :full_name, role = :role,
                                       student_number = :student_number, grade = :grade, 
                                       class_number = :class_number, number_in_class = :number_in_class
                                   WHERE user_id = :user_id"""),
                            {
                                "user_id": user_id,
                                "full_name": full_name,
                                "role": role,
                                "student_number": student_number,
                                "grade": grade,
                                "class_number": class_number,
                                "number_in_class": number_in_class
                            }
                        )
                else:
                    # 새로 생성
                    hashed = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt())
                    db.execute(
                        text("""INSERT INTO users (user_id, password_hash, full_name, role, student_number, grade, class_number, number_in_class)
                               VALUES (:user_id, :password_hash, :full_name, :role, :student_number, :grade, :class_number, :number_in_class)"""),
                        {
                            "user_id": user_id,
                            "password_hash": hashed.decode('utf-8'),
                            "full_name": full_name,
                            "role": role,
                            "student_number": student_number,
                            "grade": grade,
                            "class_number": class_number,
                            "number_in_class": number_in_class
                        }
                    )
                
                db.commit()
                results["success"] += 1
                
            except Exception as e:
                db.rollback()
                results["errors"].append(f"Row {idx}: {str(e)}")
                results["failed"] += 1
    
    elif import_type == "subjects":
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        
        for idx, row in enumerate(rows, start=2):
            try:
                subject_code, subject_name, description = row
                
                # 필수 필드 검증
                if not subject_code or not subject_name:
                    results["errors"].append(f"Row {idx}: Missing required fields")
                    results["failed"] += 1
                    continue
                
                # 기존 과목 확인 (Raw SQL)
                result = db.execute(
                    text("SELECT * FROM subjects WHERE subject_code = :subject_code"),
                    {"subject_code": subject_code}
                )
                existing_subject = result.fetchone()
                
                if existing_subject:
                    # 업데이트
                    db.execute(
                        text("""UPDATE subjects 
                               SET subject_name = :subject_name, description = :description
                               WHERE subject_code = :subject_code"""),
                        {
                            "subject_code": subject_code,
                            "subject_name": subject_name,
                            "description": description
                        }
                    )
                else:
                    # 새로 생성
                    db.execute(
                        text("""INSERT INTO subjects (subject_code, subject_name, description)
                               VALUES (:subject_code, :subject_name, :description)"""),
                        {
                            "subject_code": subject_code,
                            "subject_name": subject_name,
                            "description": description
                        }
                    )
                
                db.commit()
                results["success"] += 1
                
            except Exception as e:
                db.rollback()
                results["errors"].append(f"Row {idx}: {str(e)}")
                results["failed"] += 1
    
    else:
        raise HTTPException(status_code=400, detail="Invalid import type")
    
    return results


# ==================== 사용자 일괄 삭제 ====================
@app.post("/api/admin/users/bulk-delete")
def bulk_delete_users(
    user_ids: List[str] = Body(...),
    current_user = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Bulk delete users (admin only)"""
    
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin only")
    
    deleted_count = 0
    
    for user_id in user_ids:
        try:
            db.execute(
                text("DELETE FROM users WHERE user_id = :user_id"),
                {"user_id": user_id}
            )
            db.commit()
            deleted_count += 1
        except Exception as e:
            db.rollback()
            continue
    
    return {"message": f"{deleted_count} users deleted"}


# ===================== 교사용 api  ====================
@app.post("/api/teacher/import-activity")
async def import_activity_records(
    file: UploadFile = File(...),
    subject_id: int = Form(...),
    grade: int = Form(...),
    class_number: int = Form(...),
    school_year: int = Form(2025),
    current_user = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    활동 기록 임포트 (자율/진로/동아리)
    
    Args:
        file: Excel 파일
        subject_id: 과목 ID (자율활동=6, 진로활동=7, 동아리활동=8)
        grade: 학년
        class_number: 반
        school_year: 학년도 (기본 2025)
    
    Returns:
        {"success": int, "failed": int, "errors": []}
    """
    
    # 과목 정보 확인
    subject = db.execute(
        text("SELECT subject_name, subject_code FROM subjects WHERE id = :id"),
        {"id": subject_id}
    ).fetchone()
    
    if not subject:
        raise HTTPException(status_code=404, detail="과목을 찾을 수 없습니다")
    
    subject_code = subject.subject_code
    
    # 파일 읽기
    contents = await file.read()
    wb = load_workbook(BytesIO(contents))
    ws = wb.active
    
    results = {
        "success": 0,
        "failed": 0,
        "errors": []
    }
    
    # 헤더 행 찾기 (번호, 성명이 있는 행)
    header_row = None
    header_col_start = 1  # 기본 A열부터
    
    for row_idx in range(1, 20):
        for col_idx in range(1, 10):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value and '번호' in str(cell_value):
                header_row = row_idx
                header_col_start = col_idx
                break
        if header_row:
            break
    
    if not header_row:
        raise HTTPException(status_code=400, detail="헤더를 찾을 수 없습니다. '번호' 컬럼이 있는지 확인하세요.")
    
    print(f"[Import Activity] 헤더 행: {header_row}, 시작 열: {header_col_start}")
    print(f"[Import Activity] 과목: {subject.subject_name} ({subject_code})")
    
    # 데이터 처리
    for row_idx in range(header_row + 1, ws.max_row + 1):
        try:
            # 동아리활동
            if subject_code == 'CLUB':
                # 컬럼: 번호(B), 성명(C), 부서구분(D), 부서명(E), 부서별이수시간(F), 학생부이수시간(G), 특기사항(H)
                number = ws.cell(row=row_idx, column=header_col_start).value
                name = ws.cell(row=row_idx, column=header_col_start + 1).value
                club_category = ws.cell(row=row_idx, column=header_col_start + 2).value
                club_name = ws.cell(row=row_idx, column=header_col_start + 3).value
                club_hours = ws.cell(row=row_idx, column=header_col_start + 4).value
                record_hours = ws.cell(row=row_idx, column=header_col_start + 5).value
                remarks = ws.cell(row=row_idx, column=header_col_start + 6).value
                
                # 빈 행 스킵
                if not number or not name:
                    continue
                
                number = int(float(number)) if number else None
                club_hours = float(club_hours) if club_hours else None
                record_hours = float(record_hours) if record_hours else None
                
                if not number or not name:
                    results["errors"].append(f"행 {row_idx}: 필수 필드 누락 (번호, 성명)")
                    results["failed"] += 1
                    continue
                
                # DB 저장
                db.execute(
                    text("""
                        INSERT INTO records 
                        (record_type, subject_id, grade, class_number, number_in_class, student_name, 
                         school_year, hours, remarks, club_category, club_name, club_hours, record_hours, created_by)
                        VALUES ('activity', :subject_id, :grade, :class_number, :number, :name,
                                :school_year, :record_hours, :remarks, :club_category, :club_name, :club_hours, :record_hours2, :created_by)
                        ON CONFLICT (school_year, subject_id, grade, class_number, number_in_class)
                        WHERE record_type = 'activity'
                        DO UPDATE SET
                            student_name = EXCLUDED.student_name,
                            hours = EXCLUDED.hours,
                            remarks = EXCLUDED.remarks,
                            club_category = EXCLUDED.club_category,
                            club_name = EXCLUDED.club_name,
                            club_hours = EXCLUDED.club_hours,
                            record_hours = EXCLUDED.record_hours,
                            updated_at = CURRENT_TIMESTAMP
                    """),
                    {
                        "subject_id": subject_id,
                        "grade": grade,
                        "class_number": class_number,
                        "number": number,
                        "name": str(name).strip(),
                        "school_year": school_year,
                        "record_hours": record_hours,
                        "remarks": str(remarks).strip() if remarks else None,
                        "club_category": str(club_category).strip() if club_category else None,
                        "club_name": str(club_name).strip() if club_name else None,
                        "club_hours": club_hours,
                        "record_hours2": record_hours,
                        "created_by": current_user.user_id
                    }
                )
            
            # 자율/진로활동
            else:
                # 컬럼: 번호, 성명, 이수시간, 특기사항
                # 자율: A4부터, 진로: B8부터
                number = ws.cell(row=row_idx, column=header_col_start).value
                name = ws.cell(row=row_idx, column=header_col_start + 1).value
                hours = ws.cell(row=row_idx, column=header_col_start + 2).value
                remarks = ws.cell(row=row_idx, column=header_col_start + 3).value
                
                # 빈 행 스킵
                if not number or not name:
                    continue
                
                number = int(float(number)) if number else None
                hours = float(hours) if hours else None
                
                if not number or not name:
                    results["errors"].append(f"행 {row_idx}: 필수 필드 누락 (번호, 성명)")
                    results["failed"] += 1
                    continue
                
                # DB 저장
                db.execute(
                    text("""
                        INSERT INTO records 
                        (record_type, subject_id, grade, class_number, number_in_class, student_name, 
                         school_year, hours, remarks, created_by)
                        VALUES ('activity', :subject_id, :grade, :class_number, :number, :name,
                                :school_year, :hours, :remarks, :created_by)
                        ON CONFLICT (school_year, subject_id, grade, class_number, number_in_class)
                        WHERE record_type = 'activity'
                        DO UPDATE SET
                            student_name = EXCLUDED.student_name,
                            hours = EXCLUDED.hours,
                            remarks = EXCLUDED.remarks,
                            updated_at = CURRENT_TIMESTAMP
                    """),
                    {
                        "subject_id": subject_id,
                        "grade": grade,
                        "class_number": class_number,
                        "number": number,
                        "name": str(name).strip(),
                        "school_year": school_year,
                        "hours": hours,
                        "remarks": str(remarks).strip() if remarks else None,
                        "created_by": current_user.user_id
                    }
                )
            
            db.commit()
            results["success"] += 1
            
        except Exception as e:
            db.rollback()
            print(f"[Import Activity] 행 {row_idx} 에러: {e}")
            results["errors"].append(f"행 {row_idx}: {str(e)}")
            results["failed"] += 1
    
    print(f"[Import Activity] 완료 - 성공: {results['success']}, 실패: {results['failed']}")
    return results


# ==================== 과목별 세특 임포트 API ====================

@app.post("/api/teacher/import-subject-records")
async def import_subject_records(
    file: UploadFile = File(...),
    current_user = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    과목별 세부능력 및 특기사항 임포트
    
    파일 구조 (헤더 1행):
    학년도, 학기, 학년, 학생개인번호, 과목, 과목코드, 반/번호, 성명, 
    학적변동 구분, 세부능력 및 특기사항, 영재·발명교육 기록사항
    
    Returns:
        {"success": int, "failed": int, "errors": []}
    """
    
    contents = await file.read()
    wb = load_workbook(BytesIO(contents))
    ws = wb.active
    
    results = {
        "success": 0,
        "failed": 0,
        "errors": []
    }
    
    # 헤더는 1행
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    
    for idx, row in enumerate(rows, start=2):
        try:
            if not any(row):
                continue
            
            # 11개 컬럼 (일부는 None일 수 있음)
            school_year = row[0] if len(row) > 0 else None
            semester = row[1] if len(row) > 1 else None
            grade = row[2] if len(row) > 2 else None
            student_number = row[3] if len(row) > 3 else None
            subject_name = row[4] if len(row) > 4 else None
            subject_code = row[5] if len(row) > 5 else None
            class_and_number = row[6] if len(row) > 6 else None
            student_name = row[7] if len(row) > 7 else None
            status = row[8] if len(row) > 8 else '재학'
            remarks = row[9] if len(row) > 9 else None
            gifted_education = row[10] if len(row) > 10 else None
            
            # 필수 필드 검증
            if not all([school_year, semester, grade, student_number, subject_name]):
                results["errors"].append(f"행 {idx}: 필수 필드 누락 (학년도, 학기, 학년, 학생개인번호, 과목)")
                results["failed"] += 1
                continue
            
            # 반/번호 파싱 (예: "6/1" → class_number=6, number_in_class=1)
            class_number = None
            number_in_class = None
            if class_and_number and '/' in str(class_and_number):
                parts = str(class_and_number).split('/')
                try:
                    class_number = int(parts[0].strip()) if parts[0].strip() else None
                    number_in_class = int(parts[1].strip()) if len(parts) > 1 and parts[1].strip() else None
                except:
                    pass
            
            # 과목 ID 찾기 (과목코드 또는 과목명으로)
            subject = db.execute(
                text("""
                    SELECT id FROM subjects 
                    WHERE subject_code = :code OR subject_name = :name 
                    LIMIT 1
                """),
                {"code": subject_code, "name": subject_name}
            ).fetchone()
            
            if not subject:
                # 과목이 없으면 생성
                new_code = subject_code if subject_code else f"AUTO_{subject_name[:10]}"
                db.execute(
                    text("""
                        INSERT INTO subjects (subject_name, subject_code, description)
                        VALUES (:name, :code, :desc)
                        ON CONFLICT (subject_code) DO NOTHING
                    """),
                    {
                        "name": str(subject_name),
                        "code": new_code,
                        "desc": f"{subject_name} 과목"
                    }
                )
                db.commit()
                
                subject = db.execute(
                    text("SELECT id FROM subjects WHERE subject_code = :code LIMIT 1"),
                    {"code": new_code}
                ).fetchone()
            
            subject_id = subject.id if subject else None
            
            if not subject_id:
                results["errors"].append(f"행 {idx}: 과목을 찾거나 생성할 수 없습니다 ({subject_name})")
                results["failed"] += 1
                continue
            
            # DB 저장
            db.execute(
                text("""
                    INSERT INTO records 
                    (record_type, school_year, semester, grade, subject_id, student_number, student_name,
                     class_and_number, class_number, number_in_class, subject_name, subject_code,
                     status, content, gifted_education, created_by)
                    VALUES ('subject', :school_year, :semester, :grade, :subject_id, :student_number, :student_name,
                            :class_and_number, :class_number, :number_in_class, :subject_name, :subject_code,
                            :status, :content, :gifted_education, :created_by)
                    ON CONFLICT (school_year, semester, grade, subject_id, student_number)
                    WHERE record_type = 'subject' AND student_number IS NOT NULL
                    DO UPDATE SET
                        student_name = EXCLUDED.student_name,
                        class_and_number = EXCLUDED.class_and_number,
                        class_number = EXCLUDED.class_number,
                        number_in_class = EXCLUDED.number_in_class,
                        status = EXCLUDED.status,
                        content = EXCLUDED.content,
                        gifted_education = EXCLUDED.gifted_education,
                        updated_at = CURRENT_TIMESTAMP
                """),
                {
                    "school_year": int(school_year),
                    "semester": int(semester),
                    "grade": int(grade),
                    "subject_id": subject_id,
                    "student_number": str(student_number),
                    "student_name": str(student_name),
                    "class_and_number": str(class_and_number) if class_and_number else None,
                    "class_number": class_number,
                    "number_in_class": number_in_class,
                    "subject_name": str(subject_name),
                    "subject_code": str(subject_code) if subject_code else None,
                    "status": str(status) if status else '재학',
                    "content": str(remarks) if remarks else None,
                    "gifted_education": str(gifted_education) if gifted_education else None,
                    "created_by": current_user.user_id
                }
            )
            
            db.commit()
            results["success"] += 1
            
        except Exception as e:
            db.rollback()
            print(f"[Import Subject] 행 {idx} 에러: {e}")
            results["errors"].append(f"행 {idx}: {str(e)}")
            results["failed"] += 1
    
    print(f"[Import Subject] 완료 - 성공: {results['success']}, 실패: {results['failed']}")
    return results


# ==================== 활동 기록 조회 API ====================

@app.get("/api/teacher/activity-records")
async def get_activity_records(
    subject_id: int,
    grade: int,
    class_number: int,
    school_year: int = 2025,
    current_user = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    활동 기록 조회 (학급별)
    
    Args:
        subject_id: 과목 ID (자율=6, 진로=7, 동아리=8)
        grade: 학년
        class_number: 반
        school_year: 학년도
    
    Returns:
        List of activity records
    """
    
    records = db.execute(
        text("""
            SELECT 
                r.*,
                s.subject_name,
                s.subject_code
            FROM records r
            JOIN subjects s ON r.subject_id = s.id
            WHERE r.record_type = 'activity'
              AND r.subject_id = :subject_id 
              AND r.grade = :grade 
              AND r.class_number = :class_number
              AND r.school_year = :school_year
            ORDER BY r.number_in_class
        """),
        {
            "subject_id": subject_id,
            "grade": grade,
            "class_number": class_number,
            "school_year": school_year
        }
    ).fetchall()
    
    return [dict(row._mapping) for row in records]


# ==================== 과목별 세특 조회 API ====================

@app.get("/api/teacher/subject-records")
async def get_subject_records(
    subject_id: int,
    school_year: int,
    semester: int,
    grade: int,
    class_number: Optional[int] = None,
    current_user = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    과목별 세특 조회
    
    Args:
        subject_id: 과목 ID
        school_year: 학년도
        semester: 학기
        grade: 학년
        class_number: 반 (선택, 없으면 전체 학년)
    
    Returns:
        List of subject records
    """
    
    query = """
        SELECT 
            r.*,
            s.subject_name,
            s.subject_code
        FROM records r
        JOIN subjects s ON r.subject_id = s.id
        WHERE r.record_type = 'subject'
          AND r.subject_id = :subject_id 
          AND r.school_year = :school_year
          AND r.semester = :semester
          AND r.grade = :grade
    """
    
    params = {
        "subject_id": subject_id,
        "school_year": school_year,
        "semester": semester,
        "grade": grade
    }
    
    if class_number:
        query += " AND r.class_number = :class_number"
        params["class_number"] = class_number
    
    query += " ORDER BY r.class_number, r.number_in_class"
    
    records = db.execute(text(query), params).fetchall()
    
    return [dict(row._mapping) for row in records]


# ==================== 활동 과목 목록 조회 ====================

@app.get("/api/teacher/activity-subjects")
async def get_activity_subjects(
    current_user = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    활동 과목 목록 조회 (자율, 진로, 동아리)
    
    Returns:
        List of activity subjects
    """
    
    subjects = db.execute(
        text("""
            SELECT id, subject_name, subject_code, description
            FROM subjects
            WHERE subject_code IN ('AUTO', 'CAREER', 'CLUB')
            ORDER BY 
                CASE subject_code
                    WHEN 'AUTO' THEN 1
                    WHEN 'CAREER' THEN 2
                    WHEN 'CLUB' THEN 3
                END
        """)
    ).fetchall()
    
    return [dict(row._mapping) for row in subjects]


# ==================== Pydantic Models 추가 ====================

class TeacherAssignmentCreate(BaseModel):
    teacher_user_id: str
    role_type: str
    grade: Optional[int] = None
    class_number: Optional[int] = None
    subject_id: Optional[int] = None
    school_year: int = 2025


class TeacherAssignmentWithDetails(BaseModel):
    id: int
    teacher_user_id: str
    role_type: str
    grade: Optional[int]
    class_number: Optional[int]
    subject_id: Optional[int]
    school_year: int
    created_at: datetime
    teacher_name: Optional[str] = None
    subject_name: Optional[str] = None

    class Config:
        from_attributes = True


# ==================== 과목 삭제 API ====================

@app.delete("/api/subjects/{subject_id}")
async def delete_subject(
    subject_id: int,
    current_user = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """과목 삭제 (admin only)"""
    if current_user.role != 'admin':
        raise HTTPException(status_code=403, detail="Admin access required")
    
    records_count = db.execute(
        text("SELECT COUNT(*) FROM records WHERE subject_id = :subject_id"),
        {"subject_id": subject_id}
    ).scalar()
    
    if records_count > 0:
        raise HTTPException(
            status_code=400, 
            detail=f"이 과목에 {records_count}개의 기록이 있어 삭제할 수 없습니다."
        )
    
    try:
        result = db.execute(
            text("DELETE FROM subjects WHERE id = :id RETURNING id"),
            {"id": subject_id}
        )
        db.commit()
        
        if result.rowcount == 0:
            raise HTTPException(status_code=404, detail="과목을 찾을 수 없습니다.")
        
        return {"message": "과목이 삭제되었습니다."}
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=str(e))


@app.post("/api/subjects/bulk-delete")
async def bulk_delete_subjects(
    subject_ids: List[int] = Body(...),
    current_user = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """과목 일괄 삭제 (admin only)"""
    if current_user.role != 'admin':
        raise HTTPException(status_code=403, detail="Admin access required")
    
    deleted = 0
    errors = []
    
    for subject_id in subject_ids:
        try:
            records_count = db.execute(
                text("SELECT COUNT(*) FROM records WHERE subject_id = :subject_id"),
                {"subject_id": subject_id}
            ).scalar()
            
            if records_count > 0:
                errors.append(f"과목 ID {subject_id}: {records_count}개의 기록이 있어 삭제 불가")
                continue
            
            db.execute(text("DELETE FROM subjects WHERE id = :id"), {"id": subject_id})
            db.commit()
            deleted += 1
        except Exception as e:
            db.rollback()
            errors.append(f"과목 ID {subject_id}: {str(e)}")
    
    return {"deleted": deleted, "errors": errors}


# ==================== 사용자 개별 삭제 API ====================

@app.delete("/api/admin/users/{user_id}")
async def delete_user(
    user_id: str,
    current_user = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """사용자 삭제 (admin only)"""
    if current_user.role != 'admin':
        raise HTTPException(status_code=403, detail="Admin access required")
    
    if user_id == current_user.user_id:
        raise HTTPException(status_code=400, detail="자기 자신은 삭제할 수 없습니다.")
    
    try:
        result = db.execute(
            text("DELETE FROM users WHERE user_id = :user_id RETURNING user_id"),
            {"user_id": user_id}
        )
        db.commit()
        
        if result.rowcount == 0:
            raise HTTPException(status_code=404, detail="사용자를 찾을 수 없습니다.")
        
        return {"message": f"사용자 {user_id}가 삭제되었습니다."}
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=str(e))


# ==================== 교사 역할 배정 API ====================

@app.get("/api/admin/teacher-assignments")
async def get_all_teacher_assignments(
    school_year: int = 2025,
    current_user = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """모든 교사 역할 배정 조회 (admin only)"""
    if current_user.role != 'admin':
        raise HTTPException(status_code=403, detail="Admin access required")
    
    result = db.execute(
        text("""
            SELECT 
                ta.*,
                u.full_name as teacher_name,
                s.subject_name
            FROM teacher_assignments ta
            LEFT JOIN users u ON ta.teacher_user_id = u.user_id
            LEFT JOIN subjects s ON ta.subject_id = s.id
            WHERE ta.school_year = :school_year
            ORDER BY ta.role_type, ta.grade, ta.class_number
        """),
        {"school_year": school_year}
    )
    
    return [dict(row._mapping) for row in result.fetchall()]


@app.post("/api/admin/teacher-assignments")
async def create_teacher_assignment(
    assignment: TeacherAssignmentCreate,
    current_user = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """교사 역할 배정 생성 (admin only)"""
    if current_user.role != 'admin':
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # 역할별 필수 필드 검증
    if assignment.role_type in ['homeroom_teacher', 'assistant_homeroom']:
        if not assignment.grade or not assignment.class_number:
            raise HTTPException(status_code=400, detail="담임/부담임은 학년과 반이 필수입니다.")
    
    if assignment.role_type == 'subject_teacher':
        if not assignment.grade or not assignment.subject_id:
            raise HTTPException(status_code=400, detail="교과교사는 학년과 과목이 필수입니다.")
    
    if assignment.role_type in ['grade_head', 'record_manager']:
        if not assignment.grade:
            raise HTTPException(status_code=400, detail="학년부장/생기부관리자는 학년이 필수입니다.")
    
    try:
        result = db.execute(
            text("""
                INSERT INTO teacher_assignments 
                (teacher_user_id, role_type, grade, class_number, subject_id, school_year)
                VALUES (:teacher_user_id, :role_type, :grade, :class_number, :subject_id, :school_year)
                RETURNING *
            """),
            {
                "teacher_user_id": assignment.teacher_user_id,
                "role_type": assignment.role_type,
                "grade": assignment.grade,
                "class_number": assignment.class_number,
                "subject_id": assignment.subject_id,
                "school_year": assignment.school_year
            }
        )
        db.commit()
        return dict(result.fetchone()._mapping)
    except Exception as e:
        db.rollback()
        if "unique" in str(e).lower():
            raise HTTPException(status_code=400, detail="이미 동일한 역할이 배정되어 있습니다.")
        raise HTTPException(status_code=400, detail=str(e))


@app.delete("/api/admin/teacher-assignments/{assignment_id}")
async def delete_teacher_assignment(
    assignment_id: int,
    current_user = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """교사 역할 배정 삭제 (admin only)"""
    if current_user.role != 'admin':
        raise HTTPException(status_code=403, detail="Admin access required")
    
    try:
        result = db.execute(
            text("DELETE FROM teacher_assignments WHERE id = :id RETURNING id"),
            {"id": assignment_id}
        )
        db.commit()
        
        if result.rowcount == 0:
            raise HTTPException(status_code=404, detail="배정을 찾을 수 없습니다.")
        
        return {"message": "역할 배정이 삭제되었습니다."}
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=str(e))


# ==================== 교사 본인 역할 조회 API ====================

@app.get("/api/teacher/my-assignments")
async def get_my_assignments(
    school_year: int = 2025,
    current_user = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """현재 로그인한 교사의 역할 배정 조회"""
    if current_user.role not in ['teacher', 'admin']:
        raise HTTPException(status_code=403, detail="교사만 접근 가능합니다.")
    
    result = db.execute(
        text("""
            SELECT 
                ta.*,
                s.subject_name,
                s.subject_code
            FROM teacher_assignments ta
            LEFT JOIN subjects s ON ta.subject_id = s.id
            WHERE ta.teacher_user_id = :user_id
              AND ta.school_year = :school_year
            ORDER BY ta.role_type, ta.grade, ta.class_number
        """),
        {"user_id": current_user.user_id, "school_year": school_year}
    )
    
    return [dict(row._mapping) for row in result.fetchall()]


@app.get("/api/teacher/my-classes")
async def get_my_classes(
    school_year: int = 2025,
    current_user = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """교사의 담당 학급 목록 조회"""
    if current_user.role not in ['teacher', 'admin']:
        raise HTTPException(status_code=403, detail="교사만 접근 가능합니다.")
    
    result = db.execute(
        text("""
            SELECT DISTINCT grade, class_number, role_type
            FROM teacher_assignments
            WHERE teacher_user_id = :user_id
              AND school_year = :school_year
              AND class_number IS NOT NULL
            ORDER BY grade, class_number
        """),
        {"user_id": current_user.user_id, "school_year": school_year}
    )
    
    return [dict(row._mapping) for row in result.fetchall()]


@app.get("/api/teacher/my-subjects")
async def get_my_subjects(
    school_year: int = 2025,
    current_user = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """교사의 담당 과목 목록 조회"""
    if current_user.role not in ['teacher', 'admin']:
        raise HTTPException(status_code=403, detail="교사만 접근 가능합니다.")
    
    result = db.execute(
        text("""
            SELECT DISTINCT 
                s.id,
                s.subject_name,
                s.subject_code,
                ta.grade,
                ta.class_number
            FROM teacher_assignments ta
            JOIN subjects s ON ta.subject_id = s.id
            WHERE ta.teacher_user_id = :user_id
              AND ta.school_year = :school_year
            ORDER BY ta.grade, s.subject_name
        """),
        {"user_id": current_user.user_id, "school_year": school_year}
    )
    
    return [dict(row._mapping) for row in result.fetchall()]


# ==================== 권한 기반 기록 조회 API ====================

@app.get("/api/teacher/accessible-records")
async def get_accessible_records(
    school_year: int = 2025,
    semester: Optional[int] = None,
    grade: Optional[int] = None,
    class_number: Optional[int] = None,
    subject_id: Optional[int] = None,
    record_type: Optional[str] = None,
    current_user = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """교사의 권한에 따른 접근 가능한 기록 조회"""
    if current_user.role not in ['teacher', 'admin']:
        raise HTTPException(status_code=403, detail="교사만 접근 가능합니다.")
    
    # Admin은 모든 기록 접근 가능
    if current_user.role == 'admin':
        query = """
            SELECT r.*, s.subject_name, s.subject_code
            FROM records r
            JOIN subjects s ON r.subject_id = s.id
            WHERE r.school_year = :school_year
        """
        params = {"school_year": school_year}
        
        if semester:
            query += " AND r.semester = :semester"
            params["semester"] = semester
        if grade:
            query += " AND r.grade = :grade"
            params["grade"] = grade
        if class_number:
            query += " AND r.class_number = :class_number"
            params["class_number"] = class_number
        if subject_id:
            query += " AND r.subject_id = :subject_id"
            params["subject_id"] = subject_id
        if record_type:
            query += " AND r.record_type = :record_type"
            params["record_type"] = record_type
        
        query += " ORDER BY r.grade, r.class_number, r.number_in_class"
        
        result = db.execute(text(query), params)
        return [dict(row._mapping) for row in result.fetchall()]
    
    # 교사의 역할 조회
    assignments = db.execute(
        text("""
            SELECT role_type, grade, class_number, subject_id
            FROM teacher_assignments
            WHERE teacher_user_id = :user_id
              AND school_year = :school_year
        """),
        {"user_id": current_user.user_id, "school_year": school_year}
    ).fetchall()
    
    if not assignments:
        return []
    
    # 권한별 조건 구성
    conditions = []
    params = {"school_year": school_year}
    
    for idx, a in enumerate(assignments):
        role = a.role_type
        a_grade = a.grade
        a_class = a.class_number
        a_subject = a.subject_id
        
        if role in ['grade_head', 'record_manager']:
            conditions.append(f"(r.grade = :grade_{idx})")
            params[f"grade_{idx}"] = a_grade
            
        elif role in ['homeroom_teacher', 'assistant_homeroom']:
            conditions.append(f"(r.grade = :grade_{idx} AND r.class_number = :class_{idx})")
            params[f"grade_{idx}"] = a_grade
            params[f"class_{idx}"] = a_class
            
        elif role == 'subject_teacher':
            if a_class:
                conditions.append(f"(r.grade = :grade_{idx} AND r.class_number = :class_{idx} AND r.subject_id = :subject_{idx})")
                params[f"class_{idx}"] = a_class
            else:
                conditions.append(f"(r.grade = :grade_{idx} AND r.subject_id = :subject_{idx})")
            params[f"grade_{idx}"] = a_grade
            params[f"subject_{idx}"] = a_subject
    
    if not conditions:
        return []
    
    where_clause = " OR ".join(conditions)
    
    query = f"""
        SELECT r.*, s.subject_name, s.subject_code
        FROM records r
        JOIN subjects s ON r.subject_id = s.id
        WHERE r.school_year = :school_year
          AND ({where_clause})
    """
    
    if semester:
        query += " AND r.semester = :semester"
        params["semester"] = semester
    if grade:
        query += " AND r.grade = :grade_filter"
        params["grade_filter"] = grade
    if class_number:
        query += " AND r.class_number = :class_filter"
        params["class_filter"] = class_number
    if subject_id:
        query += " AND r.subject_id = :subject_filter"
        params["subject_filter"] = subject_id
    if record_type:
        query += " AND r.record_type = :record_type"
        params["record_type"] = record_type
    
    query += " ORDER BY r.grade, r.class_number, r.number_in_class"
    
    result = db.execute(text(query), params)
    return [dict(row._mapping) for row in result.fetchall()]


# ==================== 역할 배정 엑셀 임포트 ====================

@app.post("/api/admin/import-teacher-assignments")
async def import_teacher_assignments(
    file: UploadFile = File(...),
    school_year: int = Form(2025),
    current_user = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """교사 역할 배정 엑셀 임포트"""
    if current_user.role != 'admin':
        raise HTTPException(status_code=403, detail="Admin access required")
    
    contents = await file.read()
    wb = load_workbook(BytesIO(contents))
    ws = wb.active
    
    results = {"success": 0, "failed": 0, "errors": []}
    
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    
    for idx, row in enumerate(rows, start=2):
        try:
            if not any(row):
                continue
            
            teacher_id = row[0] if len(row) > 0 else None
            role_type = row[1] if len(row) > 1 else None
            grade = row[2] if len(row) > 2 else None
            class_number = row[3] if len(row) > 3 else None
            subject_code = row[4] if len(row) > 4 else None
            
            if not teacher_id or not role_type:
                results["errors"].append(f"행 {idx}: 교사ID와 역할은 필수입니다.")
                results["failed"] += 1
                continue
            
            # 과목 ID 조회
            subject_id = None
            if subject_code:
                subject = db.execute(
                    text("SELECT id FROM subjects WHERE subject_code = :code"),
                    {"code": str(subject_code).strip()}
                ).fetchone()
                if subject:
                    subject_id = subject.id
            
            grade = int(grade) if grade else None
            class_number = int(class_number) if class_number else None
            
            db.execute(
                text("""
                    INSERT INTO teacher_assignments 
                    (teacher_user_id, role_type, grade, class_number, subject_id, school_year)
                    VALUES (:teacher_user_id, :role_type, :grade, :class_number, :subject_id, :school_year)
                    ON CONFLICT DO NOTHING
                """),
                {
                    "teacher_user_id": str(teacher_id).strip(),
                    "role_type": str(role_type).strip(),
                    "grade": grade,
                    "class_number": class_number,
                    "subject_id": subject_id,
                    "school_year": school_year
                }
            )
            db.commit()
            results["success"] += 1
            
        except Exception as e:
            db.rollback()
            results["errors"].append(f"행 {idx}: {str(e)}")
            results["failed"] += 1
    
    return results


# ==================== 역할 배정 엑셀 템플릿 ====================

@app.get("/api/admin/download-template/teacher-assignments")
def download_teacher_assignments_template(
    current_user = Depends(get_current_user)
):
    """교사 역할 배정 템플릿 다운로드"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin only")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "역할배정"
    
    headers = ["교사ID", "역할코드", "학년", "반", "과목코드"]
    ws.append(headers)
    
    # 예시 데이터
    ws.append(["T0300", "homeroom_teacher", 2, 3, ""])
    ws.append(["T0301", "assistant_homeroom", 2, 3, ""])
    ws.append(["T0302", "subject_teacher", 2, 3, "ENG"])
    ws.append(["T0303", "subject_teacher", 2, "", "MATH"])
    ws.append(["T0304", "grade_head", 2, "", ""])
    ws.append(["T0305", "record_manager", 2, "", ""])
    
    # 역할 설명 시트
    ws2 = wb.create_sheet("역할설명")
    ws2.append(["역할코드", "역할명", "설명"])
    ws2.append(["homeroom_teacher", "학급담임", "지정 학급의 모든 교과 조회 + 자율/진로/동아리 편집"])
    ws2.append(["assistant_homeroom", "학급부담임", "지정 학급의 자율/진로/동아리 편집"])
    ws2.append(["subject_teacher", "교과교사", "지정 학년/학급의 담당 교과만 편집"])
    ws2.append(["grade_head", "학년부장", "지정 학년 전체 조회"])
    ws2.append(["record_manager", "생기부관리자", "지정 학년 전체 조회"])
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=teacher_assignments_template.xlsx"}
    )

# ==================== 메인 엔트리포인트 ====================

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
